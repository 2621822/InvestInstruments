"""Export instrument potentials to an Excel smart table with formatting.

Features:
 - Reads `instrument_potentials` from GorbunovInvestInstruments.db
 - Sorts rows by pricePotentialRel DESC
 - Creates Excel 'structured table' (openpyxl Table) covering full range
 - Formats pricePotentialRel column as percentage (0.00%) assuming stored as fractional (e.g. 0.75 -> 75.00%)
 - Optionally also writes JSON if requested

Usage:
  python export_potentials.py --excel potentials_export.xlsx --json potentials_export.json

If output paths not provided, defaults are used (potentials_export.xlsx / potentials_export.json).
"""
from __future__ import annotations

import argparse
import logging
import sqlite3
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import datetime as dt

DB_PATH = Path("GorbunovInvestInstruments.db")
DEFAULT_XLSX = "potentials_export.xlsx"
DEFAULT_JSON = "potentials_export.json"


def export_potentials(excel_path: str = DEFAULT_XLSX, json_path: Optional[str] = None) -> dict:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"DB not found: {DB_PATH}")
    with sqlite3.connect(DB_PATH) as conn:
        # Ensure table exists
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='instrument_potentials'")
        if cur.fetchone() is None:
            raise RuntimeError("Table instrument_potentials not found")
        df = pd.read_sql(
            "SELECT uid, ticker, computedDate, prevClose, consensusPrice, pricePotentialRel, isStale "
            "FROM instrument_potentials ORDER BY pricePotentialRel DESC", conn
        )

    # ---- Data transformations ----
    # Rank (1 = highest potential)
    df.reset_index(drop=True, inplace=True)
    df.insert(0, 'Rank', df.index + 1)

    # Parse dates so Excel can apply number formats
    try:
        df['computedDate'] = pd.to_datetime(df['computedDate'], errors='coerce').dt.date
    except Exception:
        pass
    # Write Excel first
    df.to_excel(excel_path, index=False)
    wb = load_workbook(excel_path)
    ws = wb.active
    # Add table
    last_row = ws.max_row
    last_col = ws.max_column
    last_col_letter = get_column_letter(last_col)
    table_ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName="PotentialsTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    # Column headers
    headers = [cell.value for cell in ws[1]]

    # Percentage formatting for pricePotentialRel + heatmap color scale
    try:
        pct_idx = headers.index("pricePotentialRel") + 1  # 1-based
        pct_col_letter = get_column_letter(pct_idx)
        rng = f"{pct_col_letter}2:{pct_col_letter}{last_row}"
        for r in range(2, last_row + 1):
            cell = ws[f"{pct_col_letter}{r}"]
            cell.number_format = '0.00%'
        # Heatmap (red -> yellow -> green where green = max potential)
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
        )
    except ValueError:
        logging.warning("pricePotentialRel column not found for formatting")

    # Date formatting for computedDate
    # computedDate -> DD.MM.YYYY
    def _apply_date_format(col_name: str, number_format: str):
        try:
            idx = headers.index(col_name) + 1
        except ValueError:
            return
        letter = get_column_letter(idx)
        for r in range(2, last_row + 1):
            cell = ws[f"{letter}{r}"]
            # Only apply number format if cell holds a date/datetime object (openpyxl infers from Python objects)
            cell.number_format = number_format

    _apply_date_format('computedDate', 'DD.MM.YYYY')

    ws.add_table(table)
    # Optional: auto-width rough sizing
    for column_cells in ws.columns:
        header_val = column_cells[0].value or ""
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells[:100])  # limit scan
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(10, max_len + 2), 50)
    wb.save(excel_path)
    logging.info("Excel potentials exported: %s (rows=%s)", excel_path, len(df))
    if json_path:
        df.to_json(json_path, orient='records', force_ascii=False, indent=2)
        logging.info("JSON potentials exported: %s", json_path)
    return {"rows": len(df), "excel": excel_path, "json": json_path}


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Export instrument potentials to Excel / JSON")
    p.add_argument('--excel', default=DEFAULT_XLSX, help=f'Excel output path (default {DEFAULT_XLSX})')
    p.add_argument('--json', help='Optional JSON output path')
    p.add_argument('--log-level', default='INFO', help='Logging level (default INFO)')
    return p


def main() -> None:
    args = build_arg_parser().parse_args()
    logging.basicConfig(level=getattr(logging, args.log_level.upper(), logging.INFO),
                        format='%(asctime)s | %(levelname)s | %(message)s')
    export_potentials(args.excel, args.json)


if __name__ == '__main__':  # pragma: no cover
    main()
