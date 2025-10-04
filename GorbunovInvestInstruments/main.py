"""Утилиты для работы с инвестиционными инструментами и консенсус-прогнозами.

Основные возможности:
 - Управление списком перспективных бумаг (добавление, первичное наполнение, обновление атрибутов)
 - Получение и сохранение консенсус-прогнозов и целей аналитиков через API Тинькофф Инвест
 - Историзация данных (с хранением предыдущих версий, ограничением глубины и очисткой устаревших записей)
 - Экспорт данных в Excel для последующего анализа
"""

from __future__ import annotations

import argparse
import logging
import logging.handlers
import os
import time
import random
import shlex
from math import isfinite
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Iterable, Any

import openpyxl
import requests
from requests.exceptions import SSLError as RequestsSSLError
import sqlite3
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ============================================================================
# ТЕХНИЧЕСКАЯ ЧАСТЬ / ИНИЦИАЛИЗАЦИЯ / КОНФИГУРАЦИЯ
# (инфраструктурные константы, логирование, HTTP / утилиты низкого уровня)
# ============================================================================
DB_PATH = Path("GorbunovInvestInstruments.db")

# ВАЖНО: Токен не следует жёстко прописывать в коде. Если переменная окружения отсутствует — выводим предупреждение.
TOKEN = os.getenv("TINKOFF_INVEST_TOKEN")
if not TOKEN:
	logging.warning(
		"Переменная окружения TINKOFF_INVEST_TOKEN не установлена. Запросы могут завершиться ошибкой авторизации."
	)
	TOKEN = ""  # пустая строка — явный маркер отсутствия токена

# Дополнительный fallback: если переменной окружения нет, пробуем прочитать локальный файл tinkoff_token.txt
if not TOKEN:
	token_file = Path("tinkoff_token.txt")
	if token_file.exists():
		try:
			file_token = token_file.read_text(encoding="utf-8").strip()
			if file_token:
				TOKEN = file_token
				logging.info("Токен загружен из tinkoff_token.txt (переменная окружения отсутствовала).")
		except Exception as exc:
			logging.warning("Не удалось прочитать tinkoff_token.txt: %s", exc)

API_BASE_URL = "https://invest-public-api.tbank.ru/rest/tinkoff.public.invest.api.contract.v1.InstrumentsService"
FIND_ENDPOINT = "FindInstrument"
GET_ENDPOINT = "GetInstrumentBy"
GET_FORECAST_ENDPOINT = "GetForecastBy"

SESSION = requests.Session()
# Возможность отключать проверку SSL только через переменную окружения (по умолчанию включено и так безопаснее)
SESSION.verify = os.getenv("APP_DISABLE_SSL_VERIFY", "0") not in {"1", "true", "True"}
if not SESSION.verify:
	urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
	logging.warning("SSL verify отключен (APP_DISABLE_SSL_VERIFY=1). Используйте только в отладочных целях.")

API_TIMEOUT = int(os.getenv("API_TIMEOUT", "15"))  # таймаут (сек) для HTTP запросов к API (override через переменную окружения API_TIMEOUT)
API_MAX_ATTEMPTS = int(os.getenv("API_MAX_ATTEMPTS", "3"))  # максимальное число попыток (ретраев) при ошибках запроса перед отказом (override через API_MAX_ATTEMPTS)
API_BACKOFF_BASE = float(os.getenv("API_BACKOFF_BASE", "0.5"))  # базовый интервал (сек) для экспоненциальной паузы между повторными попытками
MAX_CONSENSUS_PER_UID = int(os.getenv("CONSENSUS_MAX_PER_UID", "300"))  # максимум исторических consensus_forecasts записей на один uid (старые удаляются при очистке; override через CONSENSUS_MAX_PER_UID)
MAX_TARGETS_PER_ANALYST = int(os.getenv("CONSENSUS_MAX_TARGETS_PER_ANALYST", "100"))  # максимальное число записей consensus_targets на пару (uid, company) прежде чем лишние будут удалены (override через CONSENSUS_MAX_TARGETS_PER_ANALYST)
MAX_HISTORY_DAYS = int(os.getenv("CONSENSUS_MAX_HISTORY_DAYS", "1000"))  # возраст записей (дней), старше которого данные удаляются (0/<=0 отключает)
CONSENSUS_AUTO_FETCH = os.getenv("CONSENSUS_AUTO_FETCH", "1").lower() in {"1", "true", "yes", "y"}  # (начальное значение) включать ли авто-дозагрузку прогнозов при старте и добавлении новой бумаги – используйте is_auto_fetch_enabled()

# --- Метрики процесса (аггрегируются за время жизни) ---
METRICS: dict[str, int | float] = {
	"api_requests": 0,
	"api_failures": 0,
	"api_retries": 0,
	"forecast_404": 0,
}


def is_auto_fetch_enabled() -> bool:
	"""Вернуть актуальное состояние флага авто-дозагрузки.

	Читает переменную окружения при каждом вызове, позволяя изменять поведение на лету
	(например, через web UI без перезапуска процесса).
	"""
	return os.getenv("CONSENSUS_AUTO_FETCH", "1").lower() in {"1", "true", "yes", "y"}


def current_limits() -> dict[str, int | None]:
	"""Актуальные лимиты (динамически из переменных окружения с fallback к стартовым константам)."""

	def _int(name: str, default: int) -> int:
		try:
			return int(os.getenv(name, str(default)))
		except ValueError:
			return default

	return {
		"max_consensus_per_uid": _int("CONSENSUS_MAX_PER_UID", MAX_CONSENSUS_PER_UID),
		"max_targets_per_analyst": _int("CONSENSUS_MAX_TARGETS_PER_ANALYST", MAX_TARGETS_PER_ANALYST),
		"max_history_days": _int("CONSENSUS_MAX_HISTORY_DAYS", MAX_HISTORY_DAYS),
	}


def EnsureForecastsForMissingShares(db_path: Path, token: str, *, prune: bool = True) -> None:
	"""Проверить наличие прогнозов для всех бумаг и дозагрузить для тех, у кого их ещё нет.

	Сценарии использования:
	- Автоматический вызов при старте приложения (обеспечивает появление прогнозов для новых UID,
	  добавленных вручную или внешними скриптами напрямую в таблицу perspective_shares)
	- Может вызываться после операций массового добавления (fill-start / внешняя миграция)

	Логика:
	1. Собираем список uid из perspective_shares
	2. Для каждого uid проверяем, есть ли хотя бы одна строка в consensus_forecasts ИЛИ consensus_targets
	   (достаточно одного присутствия, чтобы считать прогнозы «загруженными»)
	3. Если прогнозов нет — запрашиваем (consensus, targets) и сохраняем
	4. В конце (если были добавления и prune=True) применяем PruneHistory с глобальными лимитами
	"""

	if not token:
		logging.debug("EnsureForecastsForMissingShares: токен отсутствует — пропуск дозагрузки.")
		return

	start_ts = time.time()
	added_for = 0
	added_details: list[tuple[str, str | None]] = []  # (uid, ticker)

	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		# Получаем только те UID, у которых НЕТ ни одной записи ни в одной таблице (эффективно)
		cursor.execute(
			"""
			SELECT ps.uid
			FROM perspective_shares ps
			WHERE NOT EXISTS (SELECT 1 FROM consensus_forecasts cf WHERE cf.uid = ps.uid)
			  AND NOT EXISTS (SELECT 1 FROM consensus_targets ct WHERE ct.uid = ps.uid)
			ORDER BY ps.uid
			"""
		)
		missing_uids = [row[0] for row in cursor.fetchall()]

	if not missing_uids:
		logging.debug("EnsureForecastsForMissingShares: все бумаги уже имеют хотя бы один прогноз (consensus или targets).")
		return

	for uid in missing_uids:
		consensus, targets = GetConsensusByUid(uid, token)
		AddConsensusForecasts(db_path, consensus)
		AddConsensusTargets(db_path, targets)
		ticker: str | None = None
		if isinstance(consensus, dict):
			ticker = consensus.get("ticker") or consensus.get("figi")
		if not ticker and targets:
			first = targets[0]
			ticker = first.get("ticker") or first.get("figi")
		if consensus or targets:
			added_for += 1
			added_details.append((uid, ticker))
		else:
			logging.debug("EnsureForecastsForMissingShares: по uid %s прогнозов сейчас нет (ответ пустой).", uid)

	if added_for:
		duration = time.time() - start_ts
		details_str = ", ".join(f"{u}:{t if t else '?'}" for u, t in added_details)
		logging.info(
			"EnsureForecastsForMissingShares: дозагружены прогнозы для %s бумаг за %.2fs (%s).",
			added_for,
			duration,
			details_str,
		)
		if prune:
			limits = current_limits()
			PruneHistory(
				db_path,
				limits["max_consensus_per_uid"],
				limits["max_targets_per_analyst"],
				max_age_days=limits["max_history_days"],
			)
	else:
		logging.debug("EnsureForecastsForMissingShares: подходящих для дозагрузки бумаг не осталось.")

def setup_logging() -> None:
	"""Инициализировать логирование (консоль + файл с ротацией). Повторный вызов безопасен.

	Структура формата: Время | Уровень | Сообщение
	Параметры управляются переменными окружения:
	- APP_LOG_LEVEL
	- APP_LOG_FILE
	"""
	if getattr(setup_logging, "_configured", False):  # уже настроено — повторную настройку пропускаем
		return
	log_level = os.getenv("APP_LOG_LEVEL", "INFO").upper()
	logger = logging.getLogger()
	logger.setLevel(log_level)
	for h in list(logger.handlers):  # удаляем предустановленные обработчики (если есть)
		logger.removeHandler(h)
	formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
	console = logging.StreamHandler()
	console.setFormatter(formatter)
	logger.addHandler(console)
	log_file = os.getenv("APP_LOG_FILE", "app.log")
	file_handler = logging.handlers.RotatingFileHandler(log_file, maxBytes=1_000_000, backupCount=3, encoding="utf-8")
	file_handler.setFormatter(formatter)
	logger.addHandler(file_handler)
	setup_logging._configured = True  # служебный флаг, чтобы не конфигурировать повторно (type: ignore)


setup_logging()

# Optional automatic history update & potentials recompute on start (controlled by env variables)
try:  # isolated so any failure does not break primary functionality
	from . import auto_update as _auto_update  # type: ignore
	_auto_update.maybe_run_on_start()
except Exception as _auto_exc:  # noqa: F841, BLE001
	logging.debug("auto_update initialization skipped: %s", _auto_exc)

# Необязательное принудительное применение UTF-8 для stdout (устранение "кракозябр" в консоли Windows)
if os.getenv("APP_FORCE_UTF8", "0") in {"1", "true", "True"}:
	try:
		import sys
		if hasattr(sys.stdout, "reconfigure"):
			sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
			logging.info("Консоль переведена в UTF-8 (APP_FORCE_UTF8=1)")
	except Exception as _enc_exc:  # noqa: F841
		logging.warning("Не удалось переключить stdout в UTF-8")


# --- API HELPERS -------------------------------------------------------------
def PostApiHeaders(token: str) -> dict[str, str]:
	"""Сформировать заголовки авторизации для POST запросов к API."""

	return {"Authorization": f"Bearer {token}"}


def _post(endpoint: str, payload: dict, token: str) -> dict | None:
	"""Отправить POST запрос к API с повторными попытками.

	Поведение повторов управляется переменными окружения:
	- API_MAX_ATTEMPTS — максимальное число попыток
	- API_BACKOFF_BASE — базовый интервал экспоненциальной задержки (backoff = base * 2^(attempt-1))

	Возврат: dict (JSON) или None при неуспехе.
	"""

	url = f"{API_BASE_URL}/{endpoint}"
	if not token:
		logging.error("Токен не задан. Запрос %s не будет выполнен.", endpoint)
		return None
	headers = PostApiHeaders(token)
	insecure_fallback_enabled = os.getenv("API_SSL_INSECURE_FALLBACK", "0").lower() in {"1", "true", "yes"}
	for attempt in range(1, API_MAX_ATTEMPTS + 1):
		try:
			response = SESSION.post(url, json=payload, headers=headers, timeout=API_TIMEOUT)
			METRICS["api_requests"] += 1
			if response.status_code == 404 and endpoint == GET_FORECAST_ENDPOINT:
				# Отсутствие консенсуса — не ошибка
				METRICS["forecast_404"] += 1
				logging.info("GetForecastBy 404 (нет данных) payload=%s", payload)
				return None
			status = response.status_code
			if 500 <= status < 600:
				raise requests.RequestException(f"Ошибка сервера {status}")
			response.raise_for_status()
			try:
				return response.json()
			except ValueError:
				logging.error("Ответ API %s не является корректным JSON", endpoint)
				return None
		except RequestsSSLError as ssl_exc:
			# Специальный fallback: при первой SSL ошибке можно попробовать отключить проверку сертификата (если разрешено).
			if insecure_fallback_enabled:
				logging.warning(
					"SSL ошибка при запросе %s: %s. Пробуем повторно с отключенной проверкой сертификата (insecure fallback).",
					endpoint,
					ssl_exc,
				)
				try:
					# Одноразовый небезопасный запрос (не меняем SESSION.verify глобально)
					with requests.Session() as tmp_sess:
						tmp_sess.verify = False
						urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
						insecure_resp = tmp_sess.post(url, json=payload, headers=headers, timeout=API_TIMEOUT)
						insecure_status = insecure_resp.status_code
						if 500 <= insecure_status < 600:
							raise requests.RequestException(f"Ошибка сервера {insecure_status}")
						insecure_resp.raise_for_status()
						try:
							return insecure_resp.json()
						except ValueError:
							logging.error("(fallback) Ответ API %s не является корректным JSON", endpoint)
							return None
				except requests.RequestException as insecure_exc:
					# Проваливаемся в общий блок повторов ниже как обычная ошибка
					last_exc = insecure_exc  # noqa: F841 - just for clarity
			# Если fallback не включен или тоже не удался – обрабатываем как обычную ошибку
			if attempt == API_MAX_ATTEMPTS:
				logging.error("Запрос к API %s не выполнен после %s попыток (SSL): %s", endpoint, attempt, ssl_exc)
				return None
			backoff = API_BACKOFF_BASE * (2 ** (attempt - 1))
			# Добавим небольшой джиттер чтобы избежать одновременных запросов при массовых операциях
			backoff += random.uniform(0, API_BACKOFF_BASE)
			logging.warning(
				"SSL сбой запроса к API %s (попытка %s/%s): %s. Повтор через %.2fс",
				endpoint,
				attempt,
				API_MAX_ATTEMPTS,
				ssl_exc,
				backoff,
			)
			time.sleep(backoff)
		except requests.RequestException as exc:
			if attempt == API_MAX_ATTEMPTS:
				logging.error("Запрос к API %s не выполнен после %s попыток: %s", endpoint, attempt, exc)
				METRICS["api_failures"] += 1
				return None
			backoff = API_BACKOFF_BASE * (2 ** (attempt - 1))
			backoff += random.uniform(0, API_BACKOFF_BASE)
			if attempt > 0:
				METRICS["api_retries"] += 1
			logging.warning(
				"Сбой запроса к API %s (попытка %s/%s): %s. Повтор через %.2fс",
				endpoint,
				attempt,
				API_MAX_ATTEMPTS,
				exc,
				backoff,
			)
			time.sleep(backoff)
	return None


def _parse_money(value: dict | int | float | None) -> float | None:
	"""Преобразовать денежное значение API формата {units, nano} в число с плавающей запятой."""

	if isinstance(value, dict):
		try:
			units = int(value.get("units", 0))
		except (TypeError, ValueError):
			units = 0
		try:
			nano = int(value.get("nano", 0))
		except (TypeError, ValueError):
			nano = 0
		return units + nano / 1_000_000_000

	if isinstance(value, (int, float)):
		return float(value)

	return None


def _float_equal(a: float | None, b: float | None, tol: float = 1e-6) -> bool:
	"""Сравнить два числа с плавающей точкой с заданной точностью (None считаются различными)."""

	if a is None and b is None:
		return True
	if a is None or b is None:
		return False
	return abs(a - b) <= tol



def _now_iso() -> str:
	"""Текущая отметка времени (UTC) в ISO формате без микросекунд."""
	return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _count_rows(conn: sqlite3.Connection, table: str, where: str | None = None, params: Iterable[Any] | None = None) -> int:
	"""Получить количество строк таблицы (с опциональным WHERE)."""
	sql = f"SELECT COUNT(*) FROM {table}"
	if where:
		sql += f" WHERE {where}"
	cur = conn.execute(sql, tuple(params) if params else ())
	return int(cur.fetchone()[0])


def _is_older(dt_str: str | None, cutoff: datetime) -> bool:
	"""Вернуть True, если строковая дата (ISO / с суффиксом Z) строго старше указанного порога.

	При невозможности распарсить дату возвращает False (такие записи не удаляем по возрасту).
	"""
	if not dt_str:
		return False
	val = dt_str.strip()
	# Нормализуем Z -> +00:00 чтобы datetime.fromisoformat понял
	if val.endswith("Z"):
		val = val[:-1] + "+00:00"
	try:
		parsed = datetime.fromisoformat(val)
		if parsed.tzinfo is None:
			parsed = parsed.replace(tzinfo=timezone.utc)
	except Exception:
		return False
	return parsed < cutoff


# ============================================================================
# ФУНКЦИОНАЛЬНАЯ ЧАСТЬ (БИЗНЕС-ЛОГИКА API / ПОИСК ИНСТРУМЕНТОВ)
# ============================================================================
def GetUidInstrument(search_phrase: str, token: str) -> str | None:
	"""Получить UID инструмента по поисковой строке (тикер или часть названия)."""

	# Поддержка прямых идентификаторов:
	lower = search_phrase.strip().lower()
	# Явные префиксы uid:, figi:, isin: -> извлекаем значение сразу
	for prefix in ("uid:", "figi:", "isin:"):
		if lower.startswith(prefix):
			val = search_phrase.split(":",1)[1].strip()
			if prefix == "uid:":
				return val
			# Для figi / isin попробуем получить инструмент и вернуть его uid
			id_type = "INSTRUMENT_ID_TYPE_FIGI" if prefix == "figi:" else "INSTRUMENT_ID_TYPE_ISIN"
			payload = {"idType": id_type, "id": val}
			data = _post(GET_ENDPOINT, payload, token)
			inst = data.get("instrument") if data else None
			if inst and inst.get("uid"):
				return inst.get("uid")
			return None
	# Попытка распознать, если пользователь просто вставил голый UID (формат UUID v4) или FIGI (буквенно-цифровой, длина 12) / ISIN (длина 12, цифры+буквы, заканчивается цифрой контрольной)
	import re
	if re.fullmatch(r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}", search_phrase.strip()):
		return search_phrase.strip()
	# Простая эвристика для FIGI (обычно 12 символов, может содержать 0-9A-Z):
	if re.fullmatch(r"[0-9A-Z]{12}", search_phrase.strip()):
		payload = {"idType": "INSTRUMENT_ID_TYPE_FIGI", "id": search_phrase.strip()}
		data = _post(GET_ENDPOINT, payload, token)
		inst = data.get("instrument") if data else None
		if inst and inst.get("uid"):
			return inst.get("uid")
	# ISIN: 12 символов, первые 2 буквы страны + 9 символов + контрольная цифра
	if re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}[0-9]", search_phrase.strip()):
		payload = {"idType": "INSTRUMENT_ID_TYPE_ISIN", "id": search_phrase.strip()}
		data = _post(GET_ENDPOINT, payload, token)
		inst = data.get("instrument") if data else None
		if inst and inst.get("uid"):
			return inst.get("uid")

	# Первая попытка: только торгуемые инструменты (apiTradeAvailableFlag=True)
	base_payload = {
		"query": search_phrase,
		"instrumentKind": "INSTRUMENT_TYPE_SHARE",
		"apiTradeAvailableFlag": True,
	}
	data = _post(FIND_ENDPOINT, base_payload, token)
	instruments: list[dict] = []
	if data:
		instruments = data.get("instruments", []) or []

	# Fallback: если ничего не нашли (или API вернуло пусто), пробуем повторно БЕЗ apiTradeAvailableFlag
	if not instruments:
		fallback_payload = {
			"query": search_phrase,
			"instrumentKind": "INSTRUMENT_TYPE_SHARE",
		}
		fallback_data = _post(FIND_ENDPOINT, fallback_payload, token)
		if fallback_data:
			instruments = fallback_data.get("instruments", []) or []
			if instruments:
				logging.debug(
					"GetUidInstrument: найдено только после fallback без apiTradeAvailableFlag (query='%s', count=%s)",
					search_phrase,
					len(instruments),
				)

	if not instruments:
		logging.warning("Инструменты по запросу '%s' не найдены ни в основной выборке, ни в fallback", search_phrase)
		return None

	lower_query = search_phrase.lower()
	for inst in instruments:
		ticker = inst.get("ticker", "")
		name = inst.get("name", "")
		if ticker.lower() == lower_query or name.lower() == lower_query:
			return inst.get("uid")

	return instruments[0].get("uid")


def GetInstrumentByUid(uid: str, token: str) -> dict | None:
	"""Получить полное описание инструмента по UID."""

	payload = {"idType": "INSTRUMENT_ID_TYPE_UID", "id": uid}
	data = _post(GET_ENDPOINT, payload, token)
	instrument = data.get("instrument") if data else None
	if not instrument:
		logging.warning("Не удалось получить данные по uid %s", uid)
		return None

	return {
		"ticker": instrument.get("ticker", ""),
		"name": instrument.get("name", ""),
		"uid": instrument.get("uid", ""),
		"secid": instrument.get("ticker", ""),
		"isin": instrument.get("isin", ""),
		"figi": instrument.get("figi", ""),
		"classCode": instrument.get("classCode", ""),
		"instrumentType": instrument.get("instrumentType", ""),
		"assetUid": instrument.get("assetUid", ""),
	}


def GetConsensusByUid(uid: str, token: str) -> tuple[dict | None, list[dict]]:
	"""Получить консенсус-прогноз и список целей аналитиков для заданного UID инструмента."""

	payload = {"instrumentId": uid}
	data = _post(GET_FORECAST_ENDPOINT, payload, token)
	if not data:
		logging.warning("Не удалось получить консенсус прогноз по uid %s", uid)
		return None, []

	consensus = data.get("consensus")
	targets = data.get("targets", [])
	if consensus is None:
		logging.warning("Ответ API не содержит блока consensus для uid %s", uid)

	return consensus, targets


# --- DATABASE OPERATIONS -----------------------------------------------------
def initialize_database(db_path: Path) -> None:
	"""Проверить доступность файла БД SQLite (создастся автоматически при первом подключении)."""

	try:
		with sqlite3.connect(db_path) as conn:
			conn.execute("SELECT 1")
	except sqlite3.DatabaseError as exc:
		logging.error("Не удалось инициализировать базу данных: %s", exc)
		raise


def CreateTables(db_path: Path) -> None:
	"""Создать необходимые таблицы (если ещё не существуют).

	Дополнительно гарантируем наличие индекса по (uid, recommendationDate) для ускорения
	выборки последнего консенсус-прогноза.
	"""

	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		cursor.execute(
			"""CREATE TABLE IF NOT EXISTS perspective_shares (
				ticker TEXT,
				name TEXT,
				uid TEXT PRIMARY KEY,
				secid TEXT,
				isin TEXT,
				figi TEXT,
				classCode TEXT,
				instrumentType TEXT,
				assetUid TEXT
			)"""
		)
		cursor.execute(
			"""CREATE TABLE IF NOT EXISTS consensus_forecasts (
				uid TEXT PRIMARY KEY,
				ticker TEXT,
				recommendation TEXT,
				recommendationDate DATE,
				currency TEXT,
				priceConsensus REAL,
				minTarget REAL,
				maxTarget REAL
			)"""
		)
		cursor.execute(
			"""CREATE TABLE IF NOT EXISTS consensus_targets (
				uid TEXT PRIMARY KEY,
				ticker TEXT,
				company TEXT,
				recommendation TEXT,
				recommendationDate DATE,
				currency TEXT,
				targetPrice REAL,
				showName TEXT
			)"""
		)
		# Индексы для ускорения выборок по тикеру и дате
		cursor.execute(
			"CREATE INDEX IF NOT EXISTS idx_consensus_forecasts_ticker ON consensus_forecasts(ticker)"
		)
		cursor.execute(
			"CREATE INDEX IF NOT EXISTS idx_consensus_targets_ticker_date ON consensus_targets(ticker, recommendationDate)"
		)
		# Индекс для ускорения выборки последней записи по uid
		cursor.execute(
			"CREATE INDEX IF NOT EXISTS idx_consensus_forecasts_uid_date ON consensus_forecasts(uid, recommendationDate)"
		)
		conn.commit()


def migrate_schema(db_path: Path) -> None:
	"""Мигрировать схему БД для поддержки истории и нескольких аналитиков.

	Старый вариант имел PRIMARY KEY(uid) и не позволял хранить историю.
	Новый формат:
	- consensus_forecasts: id AUTOINCREMENT, несколько записей на один uid
	- consensus_targets: id AUTOINCREMENT + уникальность (uid, company, recommendationDate)
	"""
	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		# Detect old consensus_forecasts (uid PRIMARY KEY)
		cursor.execute("PRAGMA table_info(consensus_forecasts)")
		cols = cursor.fetchall()
		if cols and len(cols) == 8 and cols[0][1] == "uid":
			logging.info("Миграция consensus_forecasts -> историческая модель")
			cursor.execute(
				"""CREATE TABLE IF NOT EXISTS consensus_forecasts_new (
					id INTEGER PRIMARY KEY AUTOINCREMENT,
					uid TEXT,
					ticker TEXT,
					recommendation TEXT,
					recommendationDate DATE,
					currency TEXT,
					priceConsensus REAL,
					minTarget REAL,
					maxTarget REAL
				)"""
			)
			cursor.execute(
				"INSERT INTO consensus_forecasts_new (uid, ticker, recommendation, recommendationDate, currency, priceConsensus, minTarget, maxTarget) SELECT uid, ticker, recommendation, recommendationDate, currency, priceConsensus, minTarget, maxTarget FROM consensus_forecasts"
			)
			cursor.execute("DROP TABLE consensus_forecasts")
			cursor.execute("ALTER TABLE consensus_forecasts_new RENAME TO consensus_forecasts")
			cursor.execute(
				"CREATE INDEX IF NOT EXISTS idx_consensus_forecasts_uid_date ON consensus_forecasts(uid, recommendationDate DESC)"
			)

		# Detect old consensus_targets (uid PRIMARY KEY)
		cursor.execute("PRAGMA table_info(consensus_targets)")
		cols_t = cursor.fetchall()
		if cols_t and len(cols_t) == 8 and cols_t[0][1] == "uid":
			logging.info("Миграция consensus_targets -> поддержка нескольких аналитиков")
			cursor.execute(
				"""CREATE TABLE IF NOT EXISTS consensus_targets_new (
					id INTEGER PRIMARY KEY AUTOINCREMENT,
					uid TEXT,
					ticker TEXT,
					company TEXT,
					recommendation TEXT,
					recommendationDate DATE,
					currency TEXT,
					targetPrice REAL,
					showName TEXT
				)"""
			)
			cursor.execute(
				"INSERT INTO consensus_targets_new (uid, ticker, company, recommendation, recommendationDate, currency, targetPrice, showName) SELECT uid, ticker, company, recommendation, recommendationDate, currency, targetPrice, showName FROM consensus_targets"
			)
			cursor.execute("DROP TABLE consensus_targets")
			cursor.execute("ALTER TABLE consensus_targets_new RENAME TO consensus_targets")
			cursor.execute(
				"CREATE UNIQUE INDEX IF NOT EXISTS uq_consensus_targets_uid_company_date ON consensus_targets(uid, company, recommendationDate)"
			)
		conn.commit()


def FillingStartDdata(db_path: Path) -> None:
	"""Первичное наполнение таблицы perspective_shares базовым набором бумаг."""

	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		count_before = _count_rows(conn, "perspective_shares")
		if count_before != 0:
			logging.info(
				"Таблица perspective_shares уже содержит данные (строк: %s). Стартовое наполнение пропущено.",
				count_before,
			)
			return

		start_data = [
			("7de75794-a27f-4d81-a39b-492345813822", "Яндекс"),
			("87db07bc-0e02-4e29-90bb-05e8ef791d7b", "Т-Технологии"),
			("e6123145-9665-43e0-8413-cd61b8aa9b13", "Сбер Банк"),
			("c190ff1f-1447-4227-b543-316332699ca5", "Сбер Банк - привилегированные акции"),
			("10620843-28ce-44e8-80c2-f26ceb1bd3e1", "Полюс"),
			("02cfdf61-6298-4c0f-a9ca-9cabc82afaf3", "ЛУКОЙЛ"),
		]
		cursor.executemany("INSERT INTO perspective_shares (uid, name) VALUES (?, ?)", start_data)
		conn.commit()

		count_after = _count_rows(conn, "perspective_shares")
		logging.info("Таблица perspective_shares заполнена стартовыми бумагами. Текущих строк: %s", count_after)


def FillingSharesData(db_path: Path, token: str) -> None:
	"""Дозаполнить отсутствующие атрибуты для каждой бумаги в perspective_shares."""

	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		cursor.execute("SELECT * FROM perspective_shares")
		rows = cursor.fetchall()
		columns = [desc[0] for desc in cursor.description]

		updated_rows = 0
		for row in rows:
			data = dict(zip(columns, row))
			missing = [col for col in columns if col != "uid" and not data.get(col)]
			if not missing:
				continue
			full_data = GetInstrumentByUid(data["uid"], token)
			if not full_data:
				continue
			cursor.execute(
				"""UPDATE perspective_shares SET
						ticker=?,
						name=?,
						secid=?,
						isin=?,
						figi=?,
						classCode=?,
						instrumentType=?,
						assetUid=?
					WHERE uid=?""",
				(
					full_data["ticker"], full_data["name"], full_data["secid"], full_data["isin"], full_data["figi"], full_data["classCode"], full_data["instrumentType"], full_data["assetUid"], data["uid"],
				),
			)
			updated_rows += 1
			logging.info("Обновлены атрибуты для %s (%s)", full_data["name"], full_data["ticker"])
		conn.commit()
		total_rows = _count_rows(conn, "perspective_shares")
		logging.info(
			"Обновление атрибутов завершено. Затронуто строк: %s. Текущее количество строк: %s",
			updated_rows,
			total_rows,
		)


def AddShareData(db_path: Path, search_phrase: str, token: str) -> None:
	"""Добавить новую бумагу в perspective_shares, найдя её по поисковой строке."""

	uid = GetUidInstrument(search_phrase, token)
	if not uid:
		logging.warning("Не удалось найти бумагу по запросу '%s'", search_phrase)
		return

	full_data = GetInstrumentByUid(uid, token)
	if not full_data:
		logging.warning("Не удалось получить данные по бумаге '%s'", search_phrase)
		return

	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		cursor.execute("SELECT uid FROM perspective_shares WHERE uid = ?", (uid,))
		if cursor.fetchone():
			total_rows = _count_rows(conn, "perspective_shares")
			logging.info(
				"Данные по бумаге %s уже присутствуют в таблице. Текущее количество строк: %s",
				full_data["name"],
				total_rows,
			)
			return

		cursor.execute(
			"""INSERT INTO perspective_shares
					(ticker, name, uid, secid, isin, figi, classCode, instrumentType, assetUid)
				VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
			(
				full_data["ticker"],
				full_data["name"],
				full_data["uid"],
				full_data["secid"],
				full_data["isin"],
				full_data["figi"],
				full_data["classCode"],
				full_data["instrumentType"],
				full_data["assetUid"],
			),
		)
		conn.commit()
		cursor.execute("SELECT COUNT(*) FROM perspective_shares")
		total_rows = cursor.fetchone()[0]
		logging.info(
			"Бумага %s успешно добавлена. Текущее количество строк: %s",
			full_data["name"],
			total_rows,
		)

	# Автоматическая загрузка прогнозов сразу после добавления новой бумаги (если включено)
	if is_auto_fetch_enabled():
		if token:
			with sqlite3.connect(db_path) as conn:
				cur = conn.cursor()
				cur.execute("SELECT COUNT(*) FROM consensus_forecasts WHERE uid=?", (full_data["uid"],))
				before_cf = cur.fetchone()[0]
				cur.execute("SELECT COUNT(*) FROM consensus_targets WHERE uid=?", (full_data["uid"],))
				before_ct = cur.fetchone()[0]
			consensus, targets = GetConsensusByUid(full_data["uid"], token)
			AddConsensusForecasts(db_path, consensus)
			AddConsensusTargets(db_path, targets)
			with sqlite3.connect(db_path) as conn:
				cur = conn.cursor()
				cur.execute("SELECT COUNT(*) FROM consensus_forecasts WHERE uid=?", (full_data["uid"],))
				after_cf = cur.fetchone()[0]
				cur.execute("SELECT COUNT(*) FROM consensus_targets WHERE uid=?", (full_data["uid"],))
				after_ct = cur.fetchone()[0]
			added_cf = after_cf - before_cf
			added_ct = after_ct - before_ct
			limits = current_limits()
			PruneHistory(
				db_path,
				limits["max_consensus_per_uid"],
				limits["max_targets_per_analyst"],
				max_age_days=limits["max_history_days"],
			)
			if added_cf or added_ct:
				logging.info(
					"Автозагрузка прогнозов для %s: добавлено consensus=%s, targets=%s.",
					full_data["ticker"],
					added_cf,
					added_ct,
				)
			else:
				logging.info(
					"Автозагрузка прогнозов для %s: в источнике пока нет данных.",
					full_data["ticker"],
				)
		else:
			logging.warning(
				"Токен отсутствует — автоматическая загрузка прогнозов для новой бумаги %s пропущена.",
				full_data["ticker"],
			)
	else:
		logging.debug("CONSENSUS_AUTO_FETCH=0 — автозагрузка прогнозов при добавлении %s отключена.", full_data["ticker"])


def AddShareByIdentifier(db_path: Path, identifier: str, id_kind: str, token: str) -> None:
	"""Добавить бумагу по прямому идентификатору (UID / FIGI / ISIN).

	id_kind: one of 'UID', 'FIGI', 'ISIN'
	"""
	kind_map = {
		'UID': 'INSTRUMENT_ID_TYPE_UID',
		'FIGI': 'INSTRUMENT_ID_TYPE_FIGI',
		'ISIN': 'INSTRUMENT_ID_TYPE_ISIN',
	}
	id_type = kind_map.get(id_kind.upper())
	if not id_type:
		logging.error("Неизвестный тип идентификатора %s", id_kind)
		return
	if not token:
		logging.error("Токен отсутствует — нельзя загрузить инструмент по %s", id_kind)
		return
	data = _post(GET_ENDPOINT, {"idType": id_type, "id": identifier}, token)
	inst = data.get("instrument") if data else None
	if not inst:
		logging.warning("Инструмент по %s=%s не найден", id_kind, identifier)
		return
	full_data = {
		"ticker": inst.get("ticker", ""),
		"name": inst.get("name", ""),
		"uid": inst.get("uid", ""),
		"secid": inst.get("ticker", ""),
		"isin": inst.get("isin", ""),
		"figi": inst.get("figi", ""),
		"classCode": inst.get("classCode", ""),
		"instrumentType": inst.get("instrumentType", ""),
		"assetUid": inst.get("assetUid", ""),
	}
	if not full_data["uid"]:
		logging.warning("Ответ по %s=%s не содержит uid — пропуск", id_kind, identifier)
		return
	with sqlite3.connect(db_path) as conn:
		cur = conn.cursor()
		cur.execute("SELECT 1 FROM perspective_shares WHERE uid=?", (full_data["uid"],))
		if cur.fetchone():
			logging.info("Инструмент %s (%s) уже присутствует.", full_data["name"], full_data["ticker"])
			return
		cur.execute(
			"""INSERT INTO perspective_shares
					(ticker, name, uid, secid, isin, figi, classCode, instrumentType, assetUid)
				 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
			(
				full_data["ticker"],
				full_data["name"],
				full_data["uid"],
				full_data["secid"],
				full_data["isin"],
				full_data["figi"],
				full_data["classCode"],
				full_data["instrumentType"],
				full_data["assetUid"],
			),
		)
		conn.commit()
	logging.info("Добавлен инструмент %s (%s) по %s=%s", full_data["name"], full_data["ticker"], id_kind, identifier)
	if is_auto_fetch_enabled():
		consensus, targets = GetConsensusByUid(full_data["uid"], token)
		AddConsensusForecasts(db_path, consensus)
		AddConsensusTargets(db_path, targets)
		limits = current_limits()
		PruneHistory(
			db_path,
			limits["max_consensus_per_uid"],
			limits["max_targets_per_analyst"],
			max_age_days=limits["max_history_days"],
		)
	else:
		logging.debug("CONSENSUS_AUTO_FETCH=0 — прогнозы по %s не загружались.", full_data["ticker"])


def AddSharesBatch(db_path: Path, queries: str, token: str) -> None:
	"""Добавить несколько бумаг (разделители: запятая и/или пробелы, поддержка кавычек)."""
	if not queries:
		return
	parts: list[str] = []
	for chunk in queries.split(','):
		chunk = chunk.strip()
		if not chunk:
			continue
		for p in shlex.split(chunk):
			p = p.strip()
			if p:
				parts.append(p)
	if not parts:
		logging.warning("Batch: нет валидных запросов")
		return
	added = 0
	for q in parts:
		with sqlite3.connect(db_path) as conn:
			cur = conn.cursor()
			cur.execute("SELECT COUNT(*) FROM perspective_shares")
			before = cur.fetchone()[0]
		AddShareData(db_path, q, token)
		with sqlite3.connect(db_path) as conn:
			cur = conn.cursor()
			cur.execute("SELECT COUNT(*) FROM perspective_shares")
			after = cur.fetchone()[0]
		if after > before:
			added += 1
	logging.info("Batch добавление завершено: запросов=%s, добавлено=%s", len(parts), added)


def DeleteShareData(db_path: Path, uid: str, *, delete_forecasts: bool = True) -> bool:
    """Удалить бумагу по UID.

    Параметры:
    - delete_forecasts: также удалить связанные consensus_forecasts и consensus_targets.

    Возвращает True если бумага была удалена.
    """
    removed = False
    with sqlite3.connect(db_path) as conn:
        cur = conn.cursor()
        cur.execute("SELECT uid, ticker, name FROM perspective_shares WHERE uid=?", (uid,))
        row = cur.fetchone()
        if not row:
            logging.warning("Удаление: бумага uid=%s не найдена.", uid)
            return False
        cur.execute("DELETE FROM perspective_shares WHERE uid=?", (uid,))
        removed = cur.rowcount > 0
        if delete_forecasts:
            cur.execute("DELETE FROM consensus_forecasts WHERE uid=?", (uid,))
            cur.execute("DELETE FROM consensus_targets WHERE uid=?", (uid,))
        conn.commit()
    if removed:
        logging.info("Удалена бумага uid=%s (%s). Связанные прогнозы удалены=%s", uid, row[1], delete_forecasts)
    return removed


def AddConsensusForecasts(db_path: Path, consensus: dict | None) -> None:
	"""Сохранить консенсус-прогноз, если он отличается от последней сохранённой записи."""

	if not consensus:
		logging.info("Консенсус данные отсутствуют, сохранение пропущено.")
		return

	uid = consensus.get("uid")
	if not uid:
		logging.warning("Консенсус прогноз не содержит uid, сохранение невозможно.")
		return

	ticker = consensus.get("ticker", "")
	recommendation = consensus.get("recommendation")
	currency = consensus.get("currency")
	price_consensus = _parse_money(consensus.get("consensus"))
	min_target = _parse_money(consensus.get("minTarget"))
	max_target = _parse_money(consensus.get("maxTarget"))

	new_inserted = False
	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		cursor.execute(
			"""SELECT uid, ticker, recommendation, currency, priceConsensus, minTarget, maxTarget
		       FROM consensus_forecasts
		      WHERE uid = ?
		   ORDER BY recommendationDate DESC
		      LIMIT 1""",
			(uid,),
		)
		last_row = cursor.fetchone()

		if last_row and (
			last_row[0] == uid
			and last_row[1] == ticker
			and last_row[2] == recommendation
			and last_row[3] == currency
			and _float_equal(last_row[4], price_consensus)
			and _float_equal(last_row[5], min_target)
			and _float_equal(last_row[6], max_target)
		):
			total_rows = _count_rows(conn, "consensus_forecasts", "uid = ?", [uid])
			logging.info(
				"Прогноз по бумаге %s уже сохранен ранее. Всего записей по бумаге: %s",
				ticker,
				total_rows,
			)
			return

		recommendation_date = _now_iso()
		cursor.execute(
			"""INSERT INTO consensus_forecasts
				(uid, ticker, recommendation, recommendationDate, currency, priceConsensus, minTarget, maxTarget)
			 VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
			(
				uid,
				ticker,
				recommendation,
				recommendation_date,
				currency,
				price_consensus,
				min_target,
				max_target,
			),
		)
		conn.commit()
		new_inserted = True

		total_rows = _count_rows(conn, "consensus_forecasts", "uid = ?", [uid])

	logging.info(
		"Консенсус прогноз для %s сохранен. Текущее количество записей по бумаге: %s",
		ticker,
		total_rows,
	)
	# Авто-пересчёт потенциала только если реально добавлена новая запись
	if new_inserted:
		try:
			RecomputePotentialForUid(db_path, uid)
		except Exception as exc:  # noqa: BLE001
			logging.debug("Авто перерасчёт потенциала по %s не удался: %s", uid, exc)


def AddConsensusTargets(db_path: Path, targets: list[dict]) -> None:
	"""Сохранить цели аналитиков, пропуская уже существующие (избежание дублей)."""

	if not targets:
		logging.info("По прогнозам аналитиков данных нет.")
		return

	# Оптимизация: минимизируем количество SELECT, используя попытку вставки и обновление при расхождении.
	inserted = 0
	updated = 0
	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		for target in targets:
			uid = target.get("uid")
			ticker = target.get("ticker")
			company = target.get("company")
			recommendation = target.get("recommendation")
			recommendation_date = target.get("recommendationDate")
			currency = target.get("currency")
			target_price = _parse_money(target.get("targetPrice"))
			show_name = target.get("showName")

			if not uid or not company or not recommendation_date:
				logging.warning(
					"Пропущен прогноз аналитика из-за отсутствия ключевых полей: uid=%s, company=%s, recommendationDate=%s",
					uid,
					company,
					recommendation_date,
				)
				continue

			# Пытаемся вставить. Если запись уже есть (уникальный индекс), пропустим.
			try:
				cursor.execute(
					"""INSERT OR IGNORE INTO consensus_targets
						(uid, ticker, company, recommendation, recommendationDate, currency, targetPrice, showName)
					 VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
					(uid, ticker, company, recommendation, recommendation_date, currency, target_price, show_name),
				)
				if cursor.rowcount > 0:
					inserted += 1
					continue  # вставили — всё
			except sqlite3.DatabaseError as exc:
				logging.error("Ошибка вставки аналитического прогноза: %s", exc)
				continue

			# Запись существует — проверим, отличаются ли поля (минимальный SELECT)
			cursor.execute(
				"""SELECT recommendation, currency, targetPrice, showName
				   FROM consensus_targets
				  WHERE uid=? AND company=? AND recommendationDate=?""",
				(uid, company, recommendation_date),
			)
			row = cursor.fetchone()
			if not row:
				# Редкая гонка: запись исчезла — пробуем ещё раз вставить без IGNORE
				try:
					cursor.execute(
						"""INSERT INTO consensus_targets
							(uid, ticker, company, recommendation, recommendationDate, currency, targetPrice, showName)
						 VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
						(uid, ticker, company, recommendation, recommendation_date, currency, target_price, show_name),
					)
					inserted += 1
					continue
				except sqlite3.DatabaseError:
					continue
			old_rec, old_currency, old_price, old_show = row
			if (
				old_rec != recommendation
				or old_currency != currency
				or (old_price is None and target_price is not None)
				or (old_price is not None and not _float_equal(old_price, target_price))
				or old_show != show_name
			):
				cursor.execute(
					"""UPDATE consensus_targets
					   SET recommendation=?, currency=?, targetPrice=?, showName=?, ticker=?
					 WHERE uid=? AND company=? AND recommendationDate=?""",
					(recommendation, currency, target_price, show_name, ticker, uid, company, recommendation_date),
				)
				updated += 1

		conn.commit()
		total_rows = _count_rows(conn, "consensus_targets")

	logging.info(
		"Загрузка прогнозов аналитиков завершена. Добавлено=%s, обновлено=%s. Всего записей: %s",
		inserted,
		updated,
		total_rows,
	)


def UpdateConsensusForecasts(
	db_path: Path,
	token: str,
	*,
	uid: str | None = None,
	max_consensus: int | None = None,
	max_targets_per_analyst: int | None = None,
	max_history_days: int | None = None,
) -> None:
	"""Получить и сохранить свежие консенсус-данные, затем выполнить очистку по лимитам.

	Приоритет источников параметров (от большего к меньшему):
	1. Значения, переданные прямо в функцию (CLI аргументы)
	2. Переменные окружения
	3. Значения по умолчанию (константы в начале файла)
	"""

	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		if uid:
			cursor.execute("SELECT uid FROM perspective_shares WHERE uid = ?", (uid,))
		else:
			cursor.execute("SELECT uid FROM perspective_shares")
		rows = cursor.fetchall()

	if not rows:
		logging.info("Нет бумаг для обновления консенсус прогноза.")
		return

	for (instrument_uid,) in rows:
		consensus, targets = GetConsensusByUid(instrument_uid, token)
		AddConsensusForecasts(db_path, consensus)
		AddConsensusTargets(db_path, targets)

	eff_max_consensus = max_consensus if max_consensus is not None else MAX_CONSENSUS_PER_UID
	eff_max_targets = max_targets_per_analyst if max_targets_per_analyst is not None else MAX_TARGETS_PER_ANALYST
	eff_max_age = max_history_days if max_history_days is not None else MAX_HISTORY_DAYS
	# age pruning disabled if non-positive
	if eff_max_age is not None and eff_max_age <= 0:
		eff_max_age = None
	# После загрузки — очистка лишней истории
	PruneHistory(db_path, eff_max_consensus, eff_max_targets, max_age_days=eff_max_age)


def FillingConsensusData(db_path: Path, token: str, *, limit: int | None = None, sleep_sec: float = 0.0) -> None:
	"""Первичное массовое наполнение consensus данными (без очистки истории).

	Алгоритм:
	1. Берёт список UID из perspective_shares (опционально усечённый по limit)
	2. Для каждой бумаги запрашивает (consensus, targets)
	3. Сохраняет через AddConsensusForecasts / AddConsensusTargets

	Отличия от UpdateConsensusForecasts:
	- Не запускает PruneHistory (мы хотим сначала накопить исторический слой)
	- Более простое логирование прогресса

	Параметры:
	- limit — ограничить количество обрабатываемых бумаг
	- sleep_sec — пауза между запросами (для снижения нагрузки / обхода лимитов)
	"""

	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		cursor.execute("SELECT uid FROM perspective_shares ORDER BY uid")
		uids = [row[0] for row in cursor.fetchall()]

	if not uids:
		logging.info("Нет инструментов для первичного консенсус наполнения.")
		return

	if limit is not None and limit >= 0:
		uids = uids[:limit]

	logging.info("Старт FillingConsensusData: инструментов к обработке %s", len(uids))

	processed = 0
	for uid in uids:
		consensus, targets = GetConsensusByUid(uid, token)
		AddConsensusForecasts(db_path, consensus)
		AddConsensusTargets(db_path, targets)
		processed += 1
		if sleep_sec:
			time.sleep(sleep_sec)
		if processed % 20 == 0:  # периодический прогресс
			logging.info("FillingConsensusData прогресс: %s/%s", processed, len(uids))

	logging.info("FillingConsensusData завершено. Обработано бумаг: %s", processed)


def PruneHistory(db_path: Path, max_consensus: int, max_targets_per_analyst: int, *, max_age_days: int | None = MAX_HISTORY_DAYS) -> None:
	"""Очистить историю по лимитам и (опционально) возрасту.

	Действия:
	- Для каждого uid оставить не более max_consensus записей в consensus_forecasts
	- Для каждой пары (uid, company) оставить не более max_targets_per_analyst записей в consensus_targets
	- Если max_age_days > 0 — удалить записи старше указанного количества дней
	"""
	deleted_consensus = 0
	deleted_targets = 0
	deleted_consensus_age = 0
	deleted_targets_age = 0
	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		# --- consensus_forecasts ---
		cursor.execute("SELECT DISTINCT uid FROM consensus_forecasts")
		for (c_uid,) in cursor.fetchall():
			cursor.execute(
				"SELECT id FROM consensus_forecasts WHERE uid=? ORDER BY recommendationDate DESC, id DESC LIMIT ?", (c_uid, max_consensus)
			)
			keep_ids = {row[0] for row in cursor.fetchall()}
			cursor.execute(
				"SELECT id FROM consensus_forecasts WHERE uid=?", (c_uid,)
			)
			all_ids = {row[0] for row in cursor.fetchall()}
			remove_ids = list(all_ids - keep_ids)
			if remove_ids:
				cursor.executemany("DELETE FROM consensus_forecasts WHERE id=?", [(rid,) for rid in remove_ids])
				deleted_consensus += len(remove_ids)

		# --- consensus_targets (per uid, company) ---
		cursor.execute("SELECT DISTINCT uid, company FROM consensus_targets")
		for c_uid, company in cursor.fetchall():
			cursor.execute(
				"SELECT id FROM consensus_targets WHERE uid=? AND company=? ORDER BY recommendationDate DESC, id DESC LIMIT ?",
				(c_uid, company, max_targets_per_analyst),
			)
			keep_ids = {row[0] for row in cursor.fetchall()}
			cursor.execute(
				"SELECT id FROM consensus_targets WHERE uid=? AND company=?",
				(c_uid, company),
			)
			all_ids = {row[0] for row in cursor.fetchall()}
			remove_ids = list(all_ids - keep_ids)
			if remove_ids:
				cursor.executemany("DELETE FROM consensus_targets WHERE id=?", [(rid,) for rid in remove_ids])
				deleted_targets += len(remove_ids)

		# --- age-based pruning (optional) ---
		if max_age_days is not None and max_age_days > 0:
			cutoff_dt = datetime.now(timezone.utc) - timedelta(days=max_age_days)
			# Прямая очистка (если recommendationDate хранится в ISO) — оставляем безопасный Python отбор (уже был)
			cursor.execute("SELECT id, recommendationDate FROM consensus_forecasts")
			for _id, dt_str in cursor.fetchall():
				if _is_older(dt_str, cutoff_dt):
					cursor.execute("DELETE FROM consensus_forecasts WHERE id=?", (_id,))
					deleted_consensus_age += 1
			cursor.execute("SELECT id, recommendationDate FROM consensus_targets")
			for _id, dt_str in cursor.fetchall():
				if _is_older(dt_str, cutoff_dt):
					cursor.execute("DELETE FROM consensus_targets WHERE id=?", (_id,))
					deleted_targets_age += 1

		conn.commit()

	if any((deleted_consensus, deleted_targets, deleted_consensus_age, deleted_targets_age)):
		logging.info(
			"Очистка истории: по лимитам удалено %s consensus_forecasts, %s consensus_targets; по возрасту удалено %s + %s (дней>%s) (лимиты %s / %s)",
			deleted_consensus,
			deleted_targets,
			deleted_consensus_age,
			deleted_targets_age,
			max_age_days,
			max_consensus,
			max_targets_per_analyst,
		)
	else:
		logging.info(
			"Очистка истории: ничего не удалено (лимиты %s / %s не превышены, возрастной порог %s дней)",
			max_consensus,
			max_targets_per_analyst,
			max_age_days,
		)


def export_perspective_shares_to_excel(db_path: Path, filename: str = "perspective_shares.xlsx") -> None:
	"""Экспортировать таблицу perspective_shares в Excel."""

	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()
		cursor.execute("SELECT * FROM perspective_shares")
		rows = cursor.fetchall()
		columns = [desc[0] for desc in cursor.description]

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Perspective Shares"
	ws.append(columns)
	for row in rows:
		ws.append(list(row))

	wb.save(filename)
	logging.info("Таблица perspective_shares экспортирована в %s", filename)


def export_consensus_to_excel(db_path: Path, filename: str = "consensus_data.xlsx") -> None:
	"""Экспортировать consensus_forecasts и consensus_targets в один Excel-файл (отдельные листы)."""

	tables = [
		("consensus_forecasts", "Consensus Forecasts"),
		("consensus_targets", "Consensus Targets"),
	]

	with sqlite3.connect(db_path) as conn:
		cursor = conn.cursor()

		wb = openpyxl.Workbook()

		for index, (table_name, sheet_title) in enumerate(tables):
			cursor.execute(f"SELECT * FROM {table_name}")
			rows = cursor.fetchall()
			columns = [desc[0] for desc in cursor.description] if cursor.description else []

			ws = wb.active if index == 0 else wb.create_sheet()
			ws.title = sheet_title
			if columns:
				ws.append(columns)
			for row in rows:
				ws.append(list(row))

	wb.save(filename)
	logging.info(
		"Таблицы consensus_forecasts и consensus_targets экспортированы в %s",
		filename,
	)


# ================== ИНТЕГРАЦИЯ И РАСЧЁТ ПОТЕНЦИАЛОВ (СИНХРОННЫЙ LOADER) =====================


def _get_prev_close(secid: str) -> tuple[float | None, str | None]:
	"""Получить последнюю цену закрытия из новой таблицы moex_history_perspective_shares.

	Legacy async loader и база moex_data.db больше не используются.
	"""
	try:
		with sqlite3.connect("GorbunovInvestInstruments.db") as conn:
			cur = conn.cursor()
			cur.execute(
				"SELECT CLOSE, COALESCE(TRADE_SESSION_DATE, TRADEDATE) FROM moex_history_perspective_shares WHERE SECID=? AND CLOSE IS NOT NULL ORDER BY COALESCE(TRADE_SESSION_DATE, TRADEDATE) DESC LIMIT 1",
				(secid,),
			)
			row = cur.fetchone()
			if row and row[0] is not None:
				return float(row[0]), row[1]
	except Exception:  # noqa: BLE001
		return None, None
	return None, None


def ComputePotentials(db_path: Path, *, moex_db: Path | None = None, store: bool = True, stale_days: int = 10) -> list[dict[str, Any]]:
	"""Рассчитать и (опционально) сохранить потенциалы в instrument_potentials.

	pricePotentialRel = (consensusPrice - prevClose) / prevClose (хранится как отношение, не %).
	Если отсутствует consensusPrice или prevClose <=0 — сохраняем NULL.
	Помечаем isStale=1 если recommendationDate старше stale_days.
	TODO: учесть корпоративные действия (дивиденды/сплиты) — запланировано.
	"""
	results: list[dict[str, Any]] = []

	def _migrate_schema_if_needed() -> None:
		"""Миграция схемы instrument_potentials к формату с уникальностью (uid, computedDate) без поля computedAt.

		Legacy варианты могли содержать computedAt и не иметь computedDate. Теперь храним одну запись в день:
		PRIMARY KEY(uid, computedDate). Поле computedAt удаляется как ненужное (оставляем только дату вычисления).
		"""
		with sqlite3.connect(db_path) as conn:
			cur = conn.cursor()
			cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='instrument_potentials'")
			if not cur.fetchone():
				return
			cur.execute("PRAGMA table_info(instrument_potentials)")
			cols = [r[1] for r in cur.fetchall()]
			# Если уже нет computedAt и есть computedDate -> ничего не делаем
			if 'computedDate' in cols and 'computedAt' not in cols:
				return
			logging.info("Миграция instrument_potentials: удаляем computedAt, обеспечиваем уникальность (uid, computedDate)")
			conn.execute(
				"""CREATE TABLE instrument_potentials_new (
					uid TEXT,
					ticker TEXT,
					computedDate TEXT,
					prevClose REAL,
					consensusPrice REAL,
					pricePotentialRel REAL,
					isStale INTEGER DEFAULT 0,
					PRIMARY KEY(uid, computedDate)
				)"""
			)
			conn.execute("CREATE INDEX IF NOT EXISTS idx_instrument_potentials_rel ON instrument_potentials_new(pricePotentialRel)")
			# Если старый формат содержал computedAt используем его для выбора последней записи дня
			if 'computedAt' in cols:
				try:
					conn.execute(
						"""
						WITH ranked AS (
						  SELECT uid,ticker,substr(computedAt,1,10) AS computedDate,prevClose,consensusPrice,pricePotentialRel,isStale,
						         ROW_NUMBER() OVER (PARTITION BY uid,substr(computedAt,1,10) ORDER BY computedAt DESC) rn
						  FROM instrument_potentials
						)
						INSERT INTO instrument_potentials_new(uid,ticker,computedDate,prevClose,consensusPrice,pricePotentialRel,isStale)
						SELECT uid,ticker,computedDate,prevClose,consensusPrice,pricePotentialRel,isStale FROM ranked WHERE rn=1;
						"""
					)
				except Exception:
					# Fallback без оконных функций
					conn.execute(
						"""
						INSERT INTO instrument_potentials_new(uid,ticker,computedDate,prevClose,consensusPrice,pricePotentialRel,isStale)
						SELECT p.uid,p.ticker,substr(p.computedAt,1,10) AS computedDate,p.prevClose,p.consensusPrice,p.pricePotentialRel,p.isStale
						FROM instrument_potentials p
						WHERE p.computedAt = (
						  SELECT MAX(p2.computedAt) FROM instrument_potentials p2
						  WHERE p2.uid=p.uid AND substr(p2.computedAt,1,10)=substr(p.computedAt,1,10)
						);
						"""
					)
			else:
				# Старый формат без computedDate но возможно без computedAt? Тогда просто проецируем
				conn.execute(
					"""INSERT INTO instrument_potentials_new(uid,ticker,computedDate,prevClose,consensusPrice,pricePotentialRel,isStale)
					SELECT uid,ticker,substr(coalesce(computedDate,computedAt,datetime('now')),1,10) AS computedDate,prevClose,consensusPrice,pricePotentialRel,isStale FROM instrument_potentials"""
				)
			conn.execute("DROP TABLE instrument_potentials")
			conn.execute("ALTER TABLE instrument_potentials_new RENAME TO instrument_potentials")
			conn.commit()
			logging.info("Миграция instrument_potentials завершена")

	def _prune_duplicates() -> int:
		"""Проверка дубликатов в старой схеме (если ещё есть) – после миграции ничего не делает."""
		with sqlite3.connect(db_path) as conn:
			cur = conn.cursor()
			cur.execute("PRAGMA table_info(instrument_potentials)")
			cols = [r[1] for r in cur.fetchall()]
			if 'computedAt' in cols and 'computedDate' not in cols:
				logging.info("Обнаружен очень старый формат без computedDate – запустить миграцию")
				return 0
			return 0

	# Run migration (once per invocation if needed)
	_migrate_schema_if_needed()
	with sqlite3.connect(db_path) as conn:
		cur = conn.cursor()
		cur.execute("SELECT uid, ticker, secid FROM perspective_shares ORDER BY ticker")
		shares = cur.fetchall()
	# Подготовка таблицы
	if store:
		with sqlite3.connect(db_path) as conn:
			cur = conn.cursor()
			cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='instrument_potentials'")
			if not cur.fetchone():
				conn.execute(
					"""CREATE TABLE instrument_potentials (
						uid TEXT,
						ticker TEXT,
						computedDate TEXT,
						prevClose REAL,
						consensusPrice REAL,
						pricePotentialRel REAL,
						isStale INTEGER DEFAULT 0,
						PRIMARY KEY(uid, computedDate)
					)"""
				)
				conn.execute("CREATE INDEX IF NOT EXISTS idx_instrument_potentials_rel ON instrument_potentials(pricePotentialRel)")
			conn.commit()
	computed_date = _now_iso()[:10]
	now_utc = datetime.now(timezone.utc)
	for uid, ticker, secid in shares:
		# Последний consensus
		consensus_price: float | None = None
		rec_date: str | None = None
		with sqlite3.connect(db_path) as conn:
			c2 = conn.cursor()
			c2.execute(
				"SELECT priceConsensus, recommendationDate FROM consensus_forecasts WHERE uid=? ORDER BY recommendationDate DESC LIMIT 1",
				(uid,),
			)
			row = c2.fetchone()
			if row:
				consensus_price, rec_date = row
		prev_close, prev_trade_dt = (None, None)
		if secid:
			prev_close, prev_trade_dt = _get_prev_close(secid)
		potential = None
		if consensus_price is not None and prev_close is not None and isfinite(prev_close) and prev_close > 0:
			potential = (consensus_price - prev_close) / prev_close
		is_stale = 0
		if rec_date:
			try:
				rd = rec_date.rstrip('Z')
				if rd.endswith('+00:00') or '+' in rd:
					rd_dt = datetime.fromisoformat(rd)
				else:
					rd_dt = datetime.fromisoformat(rd)
				if rd_dt.tzinfo is None:
					rd_dt = rd_dt.replace(tzinfo=timezone.utc)
				if (now_utc - rd_dt).days > stale_days:
					is_stale = 1
			except Exception:  # noqa: BLE001
				pass
		entry = {
			"uid": uid,
			"ticker": ticker,
			"prevClose": prev_close,
			"consensusPrice": consensus_price,
			"pricePotentialRel": potential,
			"isStale": is_stale,
		}
		results.append(entry)
		if store:
			try:
				with sqlite3.connect(db_path) as conn:
					conn.execute(
						"INSERT OR REPLACE INTO instrument_potentials (uid, ticker, computedDate, prevClose, consensusPrice, pricePotentialRel, isStale) VALUES (?,?,?,?,?,?,?)",
						(uid, ticker, computed_date, prev_close, consensus_price, potential, is_stale),
					)
					conn.commit()
			except Exception as exc:  # noqa: BLE001
				logging.debug("ComputePotentials: не удалось сохранить %s: %s", ticker, exc)
	# Final prune if legacy
	if store:
		_prune_duplicates()
	logging.info("ComputePotentials: завершено для %s бумаг (дата вычисления %s)", len(results), computed_date)
	return results


def export_potentials_to_excel(db_path: Path, filename: str = "potentials.xlsx") -> None:
	with sqlite3.connect(db_path) as conn:
		cur = conn.cursor()
		# Проверим существует ли таблица
		cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='instrument_potentials'")
		if not cur.fetchone():
			logging.warning("Экспорт potentials: таблица instrument_potentials отсутствует (создан пустой файл)")
			rows = []
			headers = ["uid","ticker","computedDate","prevClose","consensusPrice","pricePotentialRel","isStale"]
		else:
			# Новая схема без computedAt
			cur.execute(
				"""
				SELECT uid, ticker, computedDate, prevClose, consensusPrice, pricePotentialRel, isStale
				FROM instrument_potentials
				ORDER BY computedDate DESC, ticker
				"""
			)
			rows = cur.fetchall()
			headers = [d[0] for d in cur.description]
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Potentials"
	ws.append(headers)
	for r in rows:
		ws.append(list(r))
	# Формат процента
	try:
		col_idx = headers.index("pricePotentialRel") + 1
		from openpyxl.styles import numbers
		for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
			for c in cell:
				if isinstance(c.value, (int, float)) and c.value is not None:
					c.number_format = '0.00%'
	except ValueError:
		pass
	wb.save(filename)
	logging.info("Экспорт potentials -> %s", filename)


def RecomputePotentialForUid(db_path: Path, uid: str) -> None:
	"""Пересчитать потенциал для одной бумаги (последний consensus + текущий prevClose).

	Создаёт/обновляет запись instrument_potentials на текущую дату (computedDate).
	"""
	# Убедимся что таблица в мигрированном формате
	with sqlite3.connect(db_path) as conn:
		cur = conn.cursor()
		cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='instrument_potentials'")
		if not cur.fetchone():
			conn.execute(
				"""CREATE TABLE instrument_potentials (
					uid TEXT,
					ticker TEXT,
					computedDate TEXT,
					prevClose REAL,
					consensusPrice REAL,
					pricePotentialRel REAL,
					isStale INTEGER DEFAULT 0,
					PRIMARY KEY(uid, computedDate)
				)"""
			)
			conn.execute("CREATE INDEX IF NOT EXISTS idx_instrument_potentials_rel ON instrument_potentials(pricePotentialRel)")
			conn.commit()
		else:
			# При необходимости миграцию выполнит ComputePotentials (оставляем как есть)
			cur.execute("PRAGMA table_info(instrument_potentials)")
			cols = [r[1] for r in cur.fetchall()]
	with sqlite3.connect(db_path) as conn:
		cur = conn.cursor()
		cur.execute("SELECT ticker, secid FROM perspective_shares WHERE uid=?", (uid,))
		row = cur.fetchone()
		if not row:
			logging.debug("RecomputePotentialForUid: uid %s отсутствует в perspective_shares", uid)
			return
		ticker, secid = row
		cur.execute(
			"SELECT priceConsensus, recommendationDate FROM consensus_forecasts WHERE uid=? ORDER BY recommendationDate DESC LIMIT 1",
			(uid,),
		)
		c_row = cur.fetchone()
	if not c_row:
		logging.debug("RecomputePotentialForUid: нет consensus для %s", uid)
		return
	cons_price, rec_date = c_row
	prev_close, _dt = _get_prev_close(secid) if secid else (None, None)
	potential = None
	if cons_price is not None and prev_close is not None and isfinite(prev_close) and prev_close > 0:
		potential = (cons_price - prev_close) / prev_close
	is_stale = 0
	if rec_date:
		try:
			rd = rec_date.rstrip('Z')
			rd_dt = datetime.fromisoformat(rd.replace('Z', '+00:00'))
			if rd_dt.tzinfo is None:
				rd_dt = rd_dt.replace(tzinfo=timezone.utc)
			if (datetime.now(timezone.utc) - rd_dt).days > 10:
				is_stale = 1
		except Exception:  # noqa: BLE001
			pass
	computed_date = _now_iso()[:10]
	with sqlite3.connect(db_path) as conn:
		conn.execute(
			"INSERT OR REPLACE INTO instrument_potentials (uid, ticker, computedDate, prevClose, consensusPrice, pricePotentialRel, isStale) VALUES (?,?,?,?,?,?,?)",
			(uid, ticker, computed_date, prev_close, cons_price, potential, is_stale),
		)
		conn.commit()
	logging.info("RecomputePotentialForUid: %s (%s) потенциал= %s%% (дата %s)", ticker, uid, f"{potential*100:.2f}" if potential is not None else 'NULL', computed_date)


# --- CLI ---------------------------------------------------------------------
def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(description="CLI-инструменты для управления списком бумаг и консенсус-прогнозами.")
	parser.add_argument("--fill-start", action="store_true", help="заполнить таблицу perspective_shares стартовыми бумагами")
	parser.add_argument("--fill-attributes", action="store_true", help="обновить отсутствующие атрибуты бумаг")
	parser.add_argument("--add-share", metavar="QUERY", help="добавить новую бумагу по поисковой фразе")
	parser.add_argument("--add-share-uid", metavar="UID", help="добавить бумагу напрямую по UID")
	parser.add_argument("--add-share-figi", metavar="FIGI", help="добавить бумагу напрямую по FIGI")
	parser.add_argument("--add-share-isin", metavar="ISIN", help="добавить бумагу напрямую по ISIN")
	parser.add_argument(
		"--search-dump",
		metavar="QUERIES",
		help="через запятую список поисковых фраз; выполнить FindInstrument (включая fallback) и сохранить результаты в search_results.json",
	)
	parser.add_argument("--export", nargs="?", const="perspective_shares.xlsx", metavar="FILENAME", help="экспортировать таблицу perspective_shares в Excel")
	parser.add_argument(
		"--export-consensus",
		nargs="?",
		const="consensus_data.xlsx",
		metavar="FILENAME",
		help="экспортировать таблицы consensus_forecasts и consensus_targets в Excel",
	)
	parser.add_argument("--update-consensus", nargs="?", const="ALL", metavar="UID", help="обновить консенсус прогнозы для всех бумаг или указанного UID")
	parser.add_argument(
		"--fill-consensus",
		action="store_true",
		help="первичное массовое наполнение консенсус прогнозов и целей аналитиков без очистки истории",
	)
	parser.add_argument(
		"--fill-consensus-limit",
		type=int,
		metavar="N",
		help="ограничить число инструментов при выполнении --fill-consensus (для теста)",
	)
	parser.add_argument(
		"--fill-consensus-sleep",
		type=float,
		metavar="SEC",
		help="пауза (секунды) между запросами при --fill-consensus для снижения нагрузки",
	)
	parser.add_argument(
		"--max-consensus",
		type=int,
		metavar="N",
		help="лимит записей consensus_forecasts на один uid (override CONSENSUS_MAX_PER_UID)",
	)
	parser.add_argument(
		"--max-targets-per-analyst",
		type=int,
		metavar="N",
		help="лимит записей consensus_targets на пару (uid, company) (override CONSENSUS_MAX_TARGETS_PER_ANALYST)",
	)
	parser.add_argument(
		"--max-history-days",
		type=int,
		metavar="DAYS",
		help="удалять записи старше N дней (override CONSENSUS_MAX_HISTORY_DAYS, 0 или отрицательное — отключить возрастное удаление)",
	)
	parser.add_argument(
		"--ensure-forecasts",
		action="store_true",
		help="проверить и дозагрузить прогнозы для бумаг, у которых нет ни одной записи (ручной запуск)",
	)
	parser.add_argument("--add-multiple", metavar="QUERIES", help="добавить несколько бумаг (через запятую или пробел)")
	parser.add_argument("--update-prices", action="store_true", help="загрузить/обновить цены (скользящее окно) для перспективных бумаг")
	parser.add_argument("--price-horizon", type=int, default=1100, help="горизонт (дней) для скользящего окна исторических цен (default 1100)")
	parser.add_argument("--compute-potential", action="store_true", help="рассчитать потенциал (consensus vs последняя цена закрытия) и сохранить")
	parser.add_argument("--export-potentials", nargs="?", const="potentials.xlsx", metavar="FILENAME", help="экспортировать таблицу instrument_potentials в Excel")
	parser.add_argument("--stats", action="store_true", help="показать метрики и размеры таблиц")
	return parser.parse_args()


def main() -> None:
	args = parse_args()

	initialize_database(DB_PATH)
	CreateTables(DB_PATH)
	migrate_schema(DB_PATH)

	if args.fill_start:
		FillingStartDdata(DB_PATH)

	if args.fill_attributes:
		FillingSharesData(DB_PATH, TOKEN)

	if args.add_share:
		AddShareData(DB_PATH, args.add_share, TOKEN)

	if getattr(args, 'add_multiple', None):
		AddSharesBatch(DB_PATH, args.add_multiple, TOKEN)

	if getattr(args, 'add_share_uid', None):
		AddShareByIdentifier(DB_PATH, args.add_share_uid, 'UID', TOKEN)

	if getattr(args, 'add_share_figi', None):
		AddShareByIdentifier(DB_PATH, args.add_share_figi, 'FIGI', TOKEN)

	if getattr(args, 'add_share_isin', None):
		AddShareByIdentifier(DB_PATH, args.add_share_isin, 'ISIN', TOKEN)

	# Отладочная опция: дамп результатов поиска по нескольким запросам
	if getattr(args, "search_dump", None):
		import json
		queries = [q.strip() for q in args.search_dump.split(",") if q.strip()]
		results: dict[str, list[dict]] = {}
		for q in queries:
			# Выполним тот же двойной поиск что и в GetUidInstrument, но сохраним сырые данные
			payload_primary = {"query": q, "instrumentKind": "INSTRUMENT_TYPE_SHARE", "apiTradeAvailableFlag": True}
			primary = _post(FIND_ENDPOINT, payload_primary, TOKEN) or {}
			instruments = primary.get("instruments") or []
			if not instruments:
				payload_fb = {"query": q, "instrumentKind": "INSTRUMENT_TYPE_SHARE"}
				fb = _post(FIND_ENDPOINT, payload_fb, TOKEN) or {}
				instruments = fb.get("instruments") or []
			results[q] = instruments
		with open("search_results.json", "w", encoding="utf-8") as f:
			json.dump(results, f, ensure_ascii=False, indent=2)
		logging.info("Результаты поиска сохранены в search_results.json для запросов: %s", ", ".join(queries))

	# После возможного добавления/обновления списка бумаг проверяем, что у всех есть прогнозы (если авто включено)
	if is_auto_fetch_enabled():
		EnsureForecastsForMissingShares(DB_PATH, TOKEN, prune=True)
	else:
		logging.debug("CONSENSUS_AUTO_FETCH=0 — пропуск EnsureForecastsForMissingShares при старте.")

	if args.export:
		export_perspective_shares_to_excel(DB_PATH, args.export)

	if args.export_consensus:
		export_consensus_to_excel(DB_PATH, args.export_consensus)

	if args.update_consensus:
		uid = None if args.update_consensus == "ALL" else args.update_consensus
		UpdateConsensusForecasts(
			DB_PATH,
			TOKEN,
			uid=uid,
			max_consensus=getattr(args, "max_consensus", None),
			max_targets_per_analyst=getattr(args, "max_targets_per_analyst", None),
			max_history_days=getattr(args, "max_history_days", None),
		)

	if args.fill_consensus:
		FillingConsensusData(
			DB_PATH,
			TOKEN,
			limit=getattr(args, "fill_consensus_limit", None),
			sleep_sec=(getattr(args, "fill_consensus_sleep", None) or 0.0),
		)

	if getattr(args, "ensure_forecasts", False):
		EnsureForecastsForMissingShares(DB_PATH, TOKEN, prune=True)

	if getattr(args, 'update_prices', False):
		# Используем синхронный загрузчик с окном
		from . import moex_history_4_perspective_shares as mh  # type: ignore
		res = mh.daily_update_all(horizon_days=getattr(args, 'price_horizon', 1100), recompute_potentials=True)
		logging.info("Цены обновлены: inserted=%s deleted_old=%s sec=%s", res.get('total_inserted'), res.get('total_deleted_old'), len(res.get('per_security', {})))
		# После обновления потенциалы уже пересчитаны (recompute_potentials=True), выведем TOP-10
		with sqlite3.connect(DB_PATH) as conn:
			cur = conn.cursor()
			try:
				cur.execute(
					"""
					SELECT p.ticker, p.pricePotentialRel, p.isStale
					FROM instrument_potentials p
					JOIN (
						SELECT uid, MAX(computedDate) AS maxd FROM instrument_potentials GROUP BY uid
					) last ON last.uid = p.uid AND last.maxd = p.computedDate
					WHERE p.pricePotentialRel IS NOT NULL
					ORDER BY p.pricePotentialRel DESC
					LIMIT 10
					"""
				)
				rows = cur.fetchall()
				if rows:
					logging.info("Top-10 potentials после price update:")
					for t, v, stale in rows:
						logging.info("  %s: %.2f%% %s", t, v*100 if v is not None else float('nan'), '(STALE)' if stale else '')
			except Exception as exc:  # noqa: BLE001
				logging.debug("Top10 price update ошибка: %s", exc)

	if getattr(args, 'compute_potential', False):
		ComputePotentials(DB_PATH, store=True)

	if getattr(args, 'export_potentials', None):
		export_potentials_to_excel(DB_PATH, args.export_potentials)

	if getattr(args, 'stats', False):
		with sqlite3.connect(DB_PATH) as conn:
			cur = conn.cursor()
			for tbl in ("perspective_shares", "consensus_forecasts", "consensus_targets", "instrument_potentials"):
				try:
					cur.execute(f"SELECT COUNT(*) FROM {tbl}")
					logging.info("%s: %s", tbl, cur.fetchone()[0])
				except Exception:
					logging.info("%s: (нет таблицы)", tbl)
			# Топ-10 по потенциалу (только последняя вычисленная дата каждой бумаги)
			try:
				cur.execute(
					"""
					SELECT p.ticker, p.pricePotentialRel, p.isStale
					FROM instrument_potentials p
					JOIN (
						SELECT uid, MAX(computedDate) AS maxd FROM instrument_potentials GROUP BY uid
					) last ON last.uid = p.uid AND last.maxd = p.computedDate
					WHERE p.pricePotentialRel IS NOT NULL
					ORDER BY p.pricePotentialRel DESC
					LIMIT 10
					"""
				)
				rows = cur.fetchall()
				if rows:
					logging.info("Top-10 potentials (ticker: value% [stale]):")
					for t, v, stale in rows:
						logging.info("  %s: %.2f%% %s", t, v * 100 if v is not None else float('nan'), '(STALE)' if stale else '')
			except Exception as exc:  # noqa: BLE001
				logging.debug("Stats top-10 potentials error: %s", exc)
		logging.info("METRICS: %s", METRICS)

	if not any((
		args.fill_start,
		args.fill_attributes,
		args.add_share,
		getattr(args, 'add_multiple', None),
		getattr(args, 'add_share_uid', None),
		getattr(args, 'add_share_figi', None),
		getattr(args, 'add_share_isin', None),
		args.export,
		args.export_consensus,
		args.update_consensus,
		args.fill_consensus,
		getattr(args, 'load_moex_history', False),
		getattr(args, 'compute_potential', False),
		getattr(args, 'export_potentials', None),
		getattr(args, 'stats', False),
	)):
		logging.info("База данных подготовлена. Дополнительные действия не выполнялись. Используйте --help для подсказки.")


if __name__ == "__main__":
	main()
