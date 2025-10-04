"""exporting.py

Изолированный экспорт (без зависимости от legacy main.py).
Экспортируем:
 - perspective_shares -> perspective_shares.xlsx
 - consensus_forecasts + consensus_targets -> consensus_data.xlsx (2 листа)
 - instrument_potentials -> potentials_export.xlsx (+ JSON при необходимости)
"""
from __future__ import annotations

import json
import logging
import sqlite3
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers

DB_PATH = Path("GorbunovInvestInstruments.db")


def export_shares(db_path: Path = DB_PATH, filename: str = "perspective_shares.xlsx") -> None:
    with sqlite3.connect(db_path) as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM perspective_shares")
        rows = cur.fetchall()
        cols = [d[0] for d in cur.description]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Shares"; ws.append(cols)
    for r in rows: ws.append(list(r))
    wb.save(filename)
    logging.info("Экспорт shares -> %s", filename)


def export_consensus(db_path: Path = DB_PATH, filename: str = "consensus_data.xlsx") -> None:
    with sqlite3.connect(db_path) as conn:
        cur = conn.cursor()
        wb = openpyxl.Workbook()
        # forecasts
        cur.execute("SELECT * FROM consensus_forecasts")
        rows_f = cur.fetchall(); cols_f = [d[0] for d in cur.description] if cur.description else []
        ws_f = wb.active; ws_f.title = "Forecasts"; ws_f.append(cols_f)
        for r in rows_f: ws_f.append(list(r))
        # targets
        cur.execute("SELECT * FROM consensus_targets")
        rows_t = cur.fetchall(); cols_t = [d[0] for d in cur.description] if cur.description else []
        ws_t = wb.create_sheet("Targets"); ws_t.append(cols_t)
        for r in rows_t: ws_t.append(list(r))
    wb.save(filename)
    logging.info("Экспорт consensus -> %s", filename)


def export_potentials(excel_name: str = "potentials_export.xlsx", json_name: str | None = None) -> dict[str, Any]:
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.cursor()
        cur.execute(
            """SELECT uid,ticker,computedDate,prevClose,consensusPrice,pricePotentialRel,isStale
                   FROM instrument_potentials ORDER BY computedDate DESC, ticker"""
        )
        rows = cur.fetchall(); cols = [d[0] for d in cur.description]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Potentials"; ws.append(cols)
    for r in rows: ws.append(list(r))
    # формат процента
    try:
        idx = cols.index("pricePotentialRel") + 1
        for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
            for c in cell:
                if isinstance(c.value, (int, float)) and c.value is not None:
                    c.number_format = numbers.FORMAT_PERCENTAGE_00
    except ValueError:
        pass
    # умная таблица
    ref = ws.dimensions
    tbl = Table(displayName="PotentialsTable", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)
    wb.save(excel_name)
    logging.info("Экспорт potentials -> %s (rows=%s)", excel_name, len(rows))
    if json_name:
        data = [dict(zip(cols, r)) for r in rows]
        with open(json_name, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        logging.info("Экспорт potentials JSON -> %s", json_name)
    return {"rows": len(rows), "excel": excel_name, "json": json_name}


def export_all(potentials_excel: str = "potentials_export.xlsx", potentials_json: str | None = None) -> None:
    export_shares(DB_PATH)
    export_consensus(DB_PATH)
    export_potentials(potentials_excel, potentials_json)


__all__ = ["export_shares", "export_consensus", "export_potentials", "export_all"]
